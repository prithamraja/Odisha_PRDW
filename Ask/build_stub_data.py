#!/usr/bin/env python3
"""
Build the Parquet stub the backend reads until the real flat data lands.

Source: the eight per-dataset dummy-data tabs of the RTGS workbook (10 rows each).
Output: <backend>/stub_data/<table>.parquet, one file per name in db_factory.TABLES.

The stub deliberately conforms to the data contract the data agent is working to
(AP_FLAT_DATA_HANDOFF.md), not to the workbook byte-for-byte:

  * column names are the demo-tab headers verbatim (duplicates get a __dupN suffix,
    same rule validate_catalog.py uses, so template SQL written against the tabs binds);
  * dates are REAL dates. Horticulture_APMIP and Fisheries store Excel day serials in
    the workbook (45294 = 2024-01-01); those columns are converted here, because the
    contract is "no Excel serials anywhere" and the catalog's date_kind for those two
    datasets has been flipped serial -> iso to match.

Cutover to real data is a DATA_DIR change — nothing here runs in production.

Usage:
    python build_stub_data.py [--xlsx <path>] [--out <dir>]
"""
from __future__ import annotations

import argparse
import re
from datetime import date, timedelta
from pathlib import Path

# Workbook tab -> flat table name (must cover db_factory.TABLES exactly).
TABS: dict[str, str] = {
    "PM-KISAN": "pm_kisan",
    "Agriculture": "agriculture",
    "Horticulture_APMIP": "horticulture_apmip",
    "Fisheries": "fisheries",
    "Sericulture": "sericulture",
    "MARKFED": "markfed",
    "RySS_Rythu_Sadhikara": "ryss",
    "Survey_Land_Records": "survey_land_records",
}

_DEFAULT_XLSX = (
    Path(__file__).resolve().parent.parent.parent / "RTGS_Data" / "RTGS _ Decision Aid.xlsx"
)
_DEFAULT_OUT = Path(__file__).resolve().parent / "stub_data"

# A column is an Excel day serial if its name reads like a date AND every value is
# a bare number inside the plausible day-serial window (1954-09-04 .. 2064-04-16).
_DATEISH = re.compile(r"(date|dob|_dt$|^dt_|expiry|issued)", re.IGNORECASE)
_SERIAL_LO, _SERIAL_HI = 20000, 60000
_EXCEL_EPOCH = date(1899, 12, 30)  # Excel's 1900 leap-year bug is already baked in


def _excel_serial_to_date(value: float) -> date:
    return _EXCEL_EPOCH + timedelta(days=int(value))


def _dedup(headers: list[str]) -> list[str]:
    """validate_catalog.py's rule, so stub columns match what the catalog was proven on."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for h in headers:
        h = str(h).strip()
        if h in seen:
            seen[h] += 1
            h = f"{h}__dup{seen[h]}"
        else:
            seen[h] = 0
        out.append(h)
    return out


def _is_serial_column(name: str, values: list) -> bool:
    if not _DATEISH.search(name):
        return False
    present = [v for v in values if v is not None]
    if not present:
        return False
    return all(
        isinstance(v, (int, float)) and not isinstance(v, bool)
        and _SERIAL_LO < v < _SERIAL_HI
        for v in present
    )


def build(xlsx: Path, out_dir: Path) -> list[tuple[str, int, int, list[str]]]:
    import openpyxl
    import pandas as pd

    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    out_dir.mkdir(parents=True, exist_ok=True)

    summary: list[tuple[str, int, int, list[str]]] = []
    for tab, table in TABS.items():
        if tab not in wb.sheetnames:
            raise SystemExit(f"ERROR: sheet '{tab}' not found in {xlsx}")
        rows = list(wb[tab].iter_rows(values_only=True))
        headers = _dedup([h for h in rows[0] if h is not None])
        width = len(headers)
        data = [
            list(r[:width]) + [None] * max(0, width - len(r))
            for r in rows[1:]
            if any(v is not None for v in r)
        ]

        columns: dict[str, list] = {
            name: [row[i] for row in data] for i, name in enumerate(headers)
        }

        converted: list[str] = []
        for name, values in columns.items():
            if _is_serial_column(name, values):
                columns[name] = [
                    None if v is None else _excel_serial_to_date(v) for v in values
                ]
                converted.append(name)

        frame = pd.DataFrame(columns, columns=headers)
        # Datetimes from the workbook are midnight-stamped days — store them as DATE
        # so ::DATE comparisons and SUBSTR(CAST(col AS TEXT), 1, 7) both behave.
        for name in frame.columns:
            if pd.api.types.is_datetime64_any_dtype(frame[name]):
                frame[name] = frame[name].dt.date
        frame.to_parquet(out_dir / f"{table}.parquet", index=False)
        summary.append((table, len(headers), len(data), converted))
    return summary


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=_DEFAULT_XLSX)
    ap.add_argument("--out", type=Path, default=_DEFAULT_OUT)
    args = ap.parse_args()

    print(f"Reading {args.xlsx}")
    print(f"Writing {args.out}")
    for table, ncols, nrows, converted in build(args.xlsx, args.out):
        note = f"  serial->date: {', '.join(converted)}" if converted else ""
        print(f"  {table:<22} {ncols:>4} columns {nrows:>4} rows{note}")


if __name__ == "__main__":
    main()
