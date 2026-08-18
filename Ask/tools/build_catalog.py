"""
Build the PR&DW catalogue modules from `AI_Chatbot_Questions.xlsx`.

    python tools/build_catalog.py [--workbook PATH] [--check]

Emits, from the signed-off workbook and nothing else:

    query_router/template_catalog.py       346 answerable templates
    query_router/unanswerable_catalog.py    30 known-unanswerable questions
    tests/data/workbook_test_report.json    the T8 execution oracle

WHY GENERATE RATHER THAN HAND-AUTHOR
    The workbook is the ratified artefact: 363 questions the operator signed off
    on 2026-08-13, every answerable one carrying SQL already executed against
    `panchayat_1.duckdb`. Hand-copying 346 SQL strings into Python would put a
    second, divergeable copy of a signed-off asset in the tree, and any later
    re-ratification would mean doing it again. Generating keeps exactly one
    source of truth and makes "the catalogue matches the workbook" a command
    (`--check`) instead of a review.

    The GENERATED FILES ARE COMMITTED. Nothing at runtime reads the xlsx, so the
    backend needs no Excel reader and the ministry can audit the finite query set
    as ordinary source.

WHAT THIS SCRIPT DECIDES (everything else is copied verbatim)
    1. The D10 geography-predicate rewrite — see `rewrite_geography`. This is the
       only edit made to any signed-off SQL string, it is mechanical, and every
       rewritten query is re-executed against the Test Report row counts by
       tests/test_catalog_execution.py.
    2. Which slots are optional — read off the SQL's own `$p IS NULL OR` guard
       rather than the Parameter Registry's per-bind-name column, because the
       guard is what actually executes. `--check` prints where the two disagree;
       there are 13, all of them questions where a normally-optional filter is
       the subject of that particular question ("which GPs planned nothing under
       {theme}" needs the theme).
    3. Scope-phrased paraphrases (decision D2), so one consolidated template is
       retrievable at district, block and GP phrasing.
    4. Which slots carry a DEFAULT — see `DEFAULTED_SLOTS`. Decision D18.P1:
       `$top_n` is a presentational LIMIT the system supplies, not something an
       officer says, so it ships optional-with-default rather than required.

DEPENDENCY
    openpyxl, which is a build-time requirement only and is deliberately NOT in
    requirements.txt. `pip install openpyxl` before running this.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

BACKEND = Path(__file__).resolve().parents[1]
REPO = BACKEND.parent
DEFAULT_WORKBOOK = REPO / "AI_Chatbot_Questions.xlsx"

TEMPLATE_OUT = BACKEND / "query_router" / "template_catalog.py"
UNANSWERABLE_OUT = BACKEND / "query_router" / "unanswerable_catalog.py"
RERANK_OUT = BACKEND / "query_router" / "rerank_context.py"
ORACLE_OUT = BACKEND / "tests" / "data" / "workbook_test_report.json"


# ── The workbook, as records ──────────────────────────────────────────────────

def read_workbook(path: Path) -> dict[str, list[dict]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: dict[str, list[dict]] = {}
    for name in wb.sheetnames:
        rows = list(wb[name].iter_rows(values_only=True))
        if not rows:
            out[name] = []
            continue
        header = [("" if h is None else str(h).strip()) for h in rows[0]]
        records = []
        for row in rows[1:]:
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            records.append(dict(zip(header, row)))
        out[name] = records
    return out


def text(value) -> str:
    return "" if value is None else str(value).strip()


# ── Geography: the D10 predicate rewrite ──────────────────────────────────────
#
# The seven views expose `gp_lgd_code` on v_activity, v_plan and v_voucher, and
# NO block or district code anywhere (verified against the built views, not
# assumed). So:
#
#   GP        binds the resolved gp_lgd_code, and the predicate moves off the
#             name onto the code. Statewide there are ~6,800 GPs and names
#             repeat freely, so a name predicate would silently aggregate every
#             namesake into one answer.
#   BLOCK /   bind the registry-validated canonical NAME, because no view
#   DISTRICT  carries their codes. Safe for the pilot and reported as a
#             view-change request; see WP3_REPORT.
#
# `v_asset` is the one view that joins geography without carrying the code. Its
# 13 GP predicates resolve through the parent activity instead, which is exactly
# equivalent — v_asset is `activity_asset ⋈ v_activity ON activity_code`, every
# v_asset row's activity_code exists in v_activity, no activity_code is
# duplicated there, and none maps to two GPs. Replace this with the plain
# `v.gp_lgd_code = $gp_name` the day v_asset exposes the column.

VIEWS_WITH_GP_CODE = {"v_activity", "v_plan", "v_voucher", "v_approval", "gram_panchayat"}
VIEWS_WITHOUT_GP_CODE = {"v_asset", "v_progress"}

_SQL_KEYWORDS = {
    "WHERE", "GROUP", "ORDER", "ON", "LEFT", "JOIN", "LIMIT", "HAVING", "UNION",
    "SELECT", "CROSS", "INNER", "RIGHT", "FULL", "USING", "AS", "AND", "OR",
}


def mask_literals(sql: str) -> str:
    """Same-length blanking of string literals and line comments.

    Offsets still index the original, so a `$name` or a `gp_name` inside a
    literal is neither counted nor rewritten. The SBM bracket makes this matter:
    86 of its queries hold regexes like 'grey ?water' and 'e-?cart'.
    """
    out = re.sub(r"'(?:[^']|'')*'", lambda m: " " * len(m.group(0)), sql)
    return re.sub(r"--[^\n]*", lambda m: " " * len(m.group(0)), out)


# ── Which geography a statement reports ONE ROW PER ───────────────────────────
#
# Read at BUILD time and emitted as `grouped_geo`, so the runtime never parses
# SQL. The echo layer needs it to tell two readings of the same unbound slot
# apart: EXP-001 with no `$gp_name` reports one row PER GP ("…incurred by each
# GP…") while PLN-001 with no `$gp_name` reports a single count over all of them
# ("…across all GPs…"). Rendering both as "all GPs" describes the first one
# wrongly — see query_router/zones.py.

_GEO_COLUMNS = {
    "district_name": re.compile(r"\bdistrict_name\b", re.IGNORECASE),
    "block_name":    re.compile(r"\bblock_name\b", re.IGNORECASE),
    "gp_name":       re.compile(r"\bgp_(?:name|lgd_code)\b", re.IGNORECASE),
}


def _split_top_level(text: str) -> list[str]:
    """Comma-split at paren depth 0, so `SUM(a, b)` stays one item."""
    out: list[str] = []
    depth = 0
    current: list[str] = []
    for ch in text:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        if ch == "," and depth == 0:
            out.append("".join(current))
            current = []
        else:
            current.append(ch)
    out.append("".join(current))
    return [piece.strip() for piece in out if piece.strip()]


def outer_select_and_group(sql: str) -> tuple[list[str], list[str]]:
    """(SELECT items, GROUP BY items) of the OUTERMOST statement.

    Depth-0 only: 145 of the geography templates carry a GROUP BY inside a
    subquery whose grain says nothing about the shape of the answer. Literals
    are masked first so a 'group by' inside an SBM regex cannot be read as a
    clause.
    """
    masked = mask_literals(sql)
    upper = masked.upper()
    depth = 0
    select_start = select_end = group_start = group_end = None
    for i, ch in enumerate(masked):
        if ch == "(":
            depth += 1
            continue
        if ch == ")":
            depth -= 1
            continue
        if depth != 0 or (i and (upper[i - 1].isalnum() or upper[i - 1] == "_")):
            continue
        if select_start is None and upper.startswith("SELECT", i):
            select_start = i + len("SELECT")
        elif select_start is not None and select_end is None and upper.startswith("FROM", i):
            select_end = i
        elif upper.startswith("GROUP BY", i):
            group_start = i + len("GROUP BY")
        elif (group_start is not None and group_end is None
              and re.match(r"(?:ORDER\s+BY|LIMIT|HAVING)\b", upper[i:])):
            group_end = i

    select = (_split_top_level(sql[select_start:select_end])
              if select_start is not None and select_end is not None else [])
    group = (_split_top_level(sql[group_start:group_end or len(sql)])
             if group_start is not None else [])
    return select, group


def grouped_geo_slots(sql: str, slots: list[dict]) -> list[str]:
    """The geography slots this statement returns one row per.

    A bare ordinal (`GROUP BY 1,2,3`, which is how 155 of these are written) is
    resolved against the SELECT list, because the ordinal alone says nothing.
    """
    names = {s["name"] for s in slots}
    select, group = outer_select_and_group(sql)
    if not group:
        return []
    expressions = []
    for item in group:
        if item.isdigit():
            index = int(item) - 1
            if 0 <= index < len(select):
                expressions.append(select[index])
        else:
            expressions.append(item)
    joined = mask_literals(" , ".join(expressions))
    return [slot for slot, pattern in _GEO_COLUMNS.items()
            if slot in names and pattern.search(joined)]


def alias_relations(sql: str) -> dict[str, str]:
    """{alias: relation} for every FROM/JOIN in the statement.

    Raises when one alias names two relations, which would make a rewrite
    ambiguous — better to stop the build than to guess which table a predicate
    sits on.
    """
    found: dict[str, str] = {}
    for m in re.finditer(r"\b(?:FROM|JOIN)\s+(\w+)\s+(?:AS\s+)?(\w+)\b",
                         mask_literals(sql), re.IGNORECASE):
        relation, alias = m.group(1), m.group(2)
        if alias.upper() in _SQL_KEYWORDS:
            continue
        if alias in found and found[alias] != relation:
            raise ValueError(
                f"alias {alias!r} names both {found[alias]!r} and {relation!r}"
            )
        found[alias] = relation
    return found


def _reflow(spacing: str, grew_by: int) -> str:
    """Give back `grew_by` of the pad spaces a longer column name just consumed.

    `spacing` is the whole `\\s*=\\s*` between column and parameter, e.g.
    `'       = '`; only the run BEFORE the `=` is padding.
    """
    pad = len(spacing) - len(spacing.lstrip(" "))
    return " " * max(1, pad - grew_by) + spacing.lstrip(" ")


def rewrite_geography(qid: str, sql: str) -> tuple[str, list[str]]:
    """Move every GP predicate off `gp_name` and onto the resolved LGD code.

    Returns (rewritten sql, notes). Block and district predicates are left
    exactly as the workbook wrote them.
    """
    aliases = alias_relations(sql)
    notes: list[str] = []
    masked = mask_literals(sql)

    # Collected first, applied right-to-left, so earlier offsets stay valid.
    edits: list[tuple[int, int, str]] = []

    # `v.gp_name IN ($gp_name, $gp_name_2)` — the GP-vs-GP comparison (TRD-010).
    for m in re.finditer(r"(\w+)\.gp_name(\s+IN\s*\(\s*)\$(\w+)(\s*,\s*)\$(\w+)(\s*\))",
                         masked, re.IGNORECASE):
        alias = m.group(1)
        relation = aliases.get(alias)
        if relation not in VIEWS_WITH_GP_CODE:
            raise ValueError(
                f"{qid}: GP list-predicate on {relation!r}, which has no gp_lgd_code"
            )
        edits.append((m.start(), m.end(),
                      f"{alias}.gp_lgd_code{m.group(2)}${m.group(3)}"
                      f"{m.group(4)}${m.group(5)}{m.group(6)}"))
        notes.append(f"{alias}.gp_name IN (…) -> {alias}.gp_lgd_code IN (…)")

    # `v.gp_name = $gp_name` — the 302-occurrence ordinary case.
    for m in re.finditer(r"(\w+)\.gp_name(\s*=\s*)\$(gp_name(?:_2)?)\b",
                         masked, re.IGNORECASE):
        if any(s <= m.start() < e for s, e, _ in edits):
            continue
        alias, spacing, param = m.group(1), m.group(2), m.group(3)
        relation = aliases.get(alias)
        if relation in VIEWS_WITH_GP_CODE:
            # The workbook pads these predicates so the `=` signs line up down
            # the WHERE clause. `gp_lgd_code` is four characters longer than
            # `gp_name`, so give four of the pad spaces back and the column
            # survives the rewrite.
            edits.append((m.start(), m.end(),
                          f"{alias}.gp_lgd_code{_reflow(spacing, 4)}${param}"))
        elif relation in VIEWS_WITHOUT_GP_CODE:
            # Equivalent, and the only form available until the view exposes the
            # column: select the parent activities of that GP, then filter this
            # view's rows to them.
            edits.append((
                m.start(), m.end(),
                f"{alias}.activity_code IN (SELECT activity_code FROM v_activity"
                f" WHERE gp_lgd_code = ${param})",
            ))
            notes.append(
                f"{relation} exposes no gp_lgd_code; the GP filter resolves "
                f"through v_activity on activity_code"
            )
        else:
            raise ValueError(
                f"{qid}: GP predicate on unknown relation {relation!r} (alias {alias!r})"
            )

    for start, end, replacement in sorted(edits, reverse=True):
        sql = sql[:start] + replacement + sql[end:]

    # Nothing may bind a GP by name after this point — the whole purpose.
    leftover = re.search(r"\w+\.gp_name\s*(?:=|\bIN\b)\s*\(?\s*\$", mask_literals(sql),
                         re.IGNORECASE)
    if leftover:
        raise ValueError(f"{qid}: a GP name predicate survived the rewrite")
    return sql, notes


# ── Slots ─────────────────────────────────────────────────────────────────────

# Workbook bind name -> entity type. Mirrors entity_validator.PARAM_ENTITY_TYPES,
# which a test asserts this agrees with; kept here so the build has no runtime
# import and the generated file can carry literal, auditable values.
PARAM_ENTITY_TYPES = {
    "date_range": "fiscal_year", "date_range_2": "fiscal_year_2",
    "district_name": "district", "block_name": "block", "block_name_2": "block_2",
    "gp_name": "gp", "gp_name_2": "gp_2",
    "focus_area": "focus_area", "theme": "theme",
    "scheme": "scheme", "scheme_2": "scheme_2", "status": "status",
    "asset_category": "asset_category", "asset_sub_category": "asset_subcategory",
    "activity_code": "activity_code", "top_n": "top_n",
    "threshold": "threshold", "amount_threshold": "amount_threshold",
    "deadline": "deadline",
}

# The slots that bind a resolved code rather than the text the user said.
CODE_BOUND = {"gp_name", "gp_name_2"}

# Slots that ship OPTIONAL WITH A DEFAULT regardless of what the SQL guard says
# (decision D18.P1). `$top_n` is the LIMIT on 91 of the 346 templates and no
# officer says "top 10" — it is a page size the system supplies. Generated as
# required, every ranking question stalls on a "how many rows?" clarification,
# which WP-4a §5.2 showed would cost the first eval run a large, uniform and
# entirely artificial accuracy loss.
#
# STRICTLY numeric-and-presentational, and the runtime asserts the same list:
# `query_router.router._DEFAULT_ENTITY_VALUES` must agree with these values
# (tests/test_param_binding.py pins it). A categorical slot or a threshold that
# changes the POPULATION under study — $threshold, $amount_threshold — must
# never appear here: guessing there silently answers a different question than
# the one asked (D18.P2; the 15 threshold-bearing templates clarify instead).
#
# The value is a string because ExtractedEntity carries strings and DuckDB casts
# at bind time, including inside LIMIT.
DEFAULTED_SLOTS = {"top_n": "10"}

# {Token} in the question text -> the slot it stands for. `.format()` is called
# on abstract_question by suggestions._chip_for, so every placeholder must be a
# real slot name or a pre-filled chip raises KeyError.
QUESTION_TOKENS = {
    "Date_Range": "date_range", "Date_Range_2": "date_range_2",
    "District": "district_name", "Block": "block_name", "Block_2": "block_name_2",
    "GP_Name": "gp_name", "GP_Name_2": "gp_name_2",
    "Focus_Area": "focus_area", "Theme": "theme",
    "Scheme": "scheme", "Scheme_2": "scheme_2", "Status": "status",
    "Asset_Category": "asset_category", "Asset_Sub_Category": "asset_sub_category",
    "Activity_Code": "activity_code", "Threshold": "threshold",
    "Amount_Threshold": "amount_threshold", "Deadline": "deadline",
}


def declared_params(row: dict) -> list[str]:
    return re.findall(r"\$(\w+)", text(row["Parameters (bind order)"]))


def is_optional(sql: str, param: str) -> bool:
    """True when the SQL itself says a missing value means "don't filter".

    The guard `$p IS NULL OR …` is the authority rather than the Parameter
    Registry's NULL-skips column, because the guard is what executes. The two
    disagree on 13 questions, every one of them a case where a normally-optional
    filter is that question's subject (SCH-006 asks which GPs planned nothing
    under a named scheme — without the scheme there is no question). The
    parenthesis is optional: TRD-012 writes the guard unbracketed.
    """
    return bool(re.search(r"\$" + re.escape(param) + r"\s+IS\s+NULL\s+OR",
                          mask_literals(sql), re.IGNORECASE))


def build_slots(row: dict, sql: str) -> list[dict]:
    slots = []
    for name in declared_params(row):
        entity_type = PARAM_ENTITY_TYPES.get(name)
        if entity_type is None:
            raise ValueError(f"{row['Question ID']}: unknown bind name ${name}")
        slot = {"name": name, "entity_type": entity_type}
        if is_optional(sql, name) or name in DEFAULTED_SLOTS:
            slot["optional"] = True
        if name in DEFAULTED_SLOTS:
            # Optional here means "the router supplies it", NOT "bind NULL" —
            # `LIMIT NULL` is unbounded, the opposite of a page size. The
            # default is declared on the slot so every serving path reads it
            # from one place; see router.slot_defaults().
            slot["default"] = DEFAULTED_SLOTS[name]
        if name in CODE_BOUND:
            slot["bind"] = "code"
        slots.append(slot)
    return slots


# ── Question text and paraphrases ─────────────────────────────────────────────

# How each placeholder reads when a question is written out in prose for the
# embedding index. The retriever strips the braces from `abstract_question`
# anyway ({district_name} -> district_name), so paraphrases are written as plain
# words instead: that is the text an officer's own question has to match.
TOKEN_PROSE = {
    "Date_Range": "a given year", "Date_Range_2": "a second year",
    "District": "a given district", "Block": "a given block",
    "Block_2": "a second block", "GP_Name": "a given gram panchayat",
    "GP_Name_2": "a second gram panchayat", "Focus_Area": "a given focus area",
    "Theme": "a given LSDG theme", "Scheme": "a given scheme",
    "Scheme_2": "a second scheme", "Status": "a given status",
    "Asset_Category": "a given asset category",
    "Asset_Sub_Category": "a given asset sub-category",
    "Activity_Code": "a given activity", "Threshold": "a given threshold",
    "Amount_Threshold": "a given amount", "Deadline": "a given deadline",
    "Subject": "a given subject",
}

# Decision D2: one consolidated template answers at every geographic scope, so
# each scope has to be RETRIEVABLE. A template gets one paraphrase per tier it
# can actually filter on — the tier NOUN is what an officer's own wording shares
# with the index ("in Khordha district", "for Andhrua panchayat").
SCOPE_NOUN = {
    "district_name": "a given district",
    "block_name": "a given block",
    "gp_name": "a given gram panchayat",
}
SCOPE_SUFFIX = {
    "district_name": "for a given district",
    "block_name": "for a given block",
    "gp_name": "for a given gram panchayat (GP)",
}

# A run of geography placeholders and whatever joins them: '{District}/{Block}',
# '{District} or {Block}', '{GP_Name}'. Replaced as a unit, so a question that
# already names its scope reads naturally in each tier's variant instead of
# collecting all three.
_GEO_TOKENS = "District|Block|Block_2|GP_Name|GP_Name_2"
_GEO_RUN = re.compile(
    r"\{(?:" + _GEO_TOKENS + r")\}(?:\s*(?:/|,|\bor\b|\band\b)\s*"
    r"\{(?:" + _GEO_TOKENS + r")\})*"
)


def to_abstract(parameterised: str, slot_names: set[str], qid: str) -> str:
    """'…in {District}…' -> '…in {district_name}…', formattable by slot name."""
    def sub(m):
        token = m.group(1)
        slot = QUESTION_TOKENS.get(token)
        if slot is None:
            raise ValueError(f"{qid}: unknown question placeholder {{{token}}}")
        if slot not in slot_names:
            # A placeholder naming a slot the template does not have would make
            # .format() raise inside suggestions._chip_for. Write it as prose.
            return TOKEN_PROSE[token]
        return "{" + slot + "}"
    return re.sub(r"\{([^}]*)\}", sub, parameterised or "")


def to_prose(question: str) -> str:
    """Every placeholder written out, for the embedding index."""
    return re.sub(
        r"\{([^}]*)\}",
        lambda m: TOKEN_PROSE.get(m.group(1), m.group(1).replace("_", " ").lower()),
        question or "",
    ).strip()


# ── Retrieval surface for the questions with no SQL ───────────────────────────
#
# A known-unanswerable row carries no parameterised SQL, so it has no
# {placeholders} and the workbook writes its parameters out as PROSE: "…under a
# given Scheme in a given GP Name during a given Plan Year?". That sentence is
# what gets embedded, and it does not retrieve. Measured (WP-4c T2c) against the
# gold question that is almost word-for-word its own:
#
#   BEN-001  "How many beneficiaries received benefits under Swachh Bharat
#             Mission in Andhrua during 2024-25?"   ->  rank 51 of 376
#
# outside the top-30 window, so the reranker never saw the entry and the
# documented refusal could not be reached at all. Same for BEN-003 (64) and
# PLN-022 (46). The cause is a plain asymmetry of index surface: a TEMPLATE
# carries 6.1 vectors on average — abstract question, example question with real
# values, one scope line per tier it can filter on (D2) — while the 13 Dropped
# rows carried exactly ONE, and that one is three-fifths filler. The Dropped
# sheet has no Parameterized or Example column to draw on, so the generator has
# to make the surface itself.
#
# IT STRIPS THE SCOPE, which is the same thing SCOPE_SUFFIX does for templates
# from the other direction: an officer names the place and the year, so the
# catalogue's copy of them adds nothing and dilutes the measure words that do the
# matching. Measured after: BEN-001 rank 1, BEN-003 rank 12, PLN-022 rank 0.
#
# A paraphrase can only ever RAISE the entry it belongs to — the retriever scores
# an entry as the MAX over its vectors — so this cannot cost any other row its
# place except by out-ranking it, which is the intended effect.

_GIVEN_PERIOD = re.compile(
    r"\s*\b(?:in|for|during|over)\s+a\s+given\s+"
    r"(?:Plan\s+Year|Date\s+Range|year)\b", re.I)
_GIVEN_GEO = re.compile(
    r"\s*\b(?:in|for|of|across|at)\s+a\s+given\s+"
    r"(?:District|Block(?:\s+2)?|GP\s+Name(?:\s+2)?)\b", re.I)


def scope_free_question(question: str) -> str | None:
    """The question with its geography and period prose removed, or None.

    None when nothing was removed — PLN-041 ("…over the last five years?") names
    no parameter, so there is no second shape of it to index.
    """
    stripped = _GIVEN_GEO.sub("", _GIVEN_PERIOD.sub("", question or ""))
    stripped = re.sub(r"\s{2,}", " ", stripped).strip().rstrip(" ,")
    if not stripped or stripped.lower() == (question or "").strip().lower():
        return None
    if not stripped.endswith(("?", ".")):
        stripped += "?"
    return stripped[0].upper() + stripped[1:]




# ── The CODE-MIXED retrieval surface (D31.5, WP-5 T4c) ────────────────────────
#
# WHAT WP-4c LEFT OPEN. Stripping the scope prose (above) moved BEN-001 from
# rank 51 to rank 4 against its own gold question and turned 0/3 replays into
# 3/3. It did nothing for BEN-003, which sat at 64 before and 64 after — because
# BEN-003's gold question is not English:
#
#   BEN-003 catalogue: "How many beneficiaries are recorded across GPs of a
#                       given Block under a given Scheme for a given Plan Year?"
#   BEN-003 gold:      "Bhubaneswar block ke GPs me kitne beneficiary hain
#                       2024-25 me?"
#
# WP-4c filed that as F2 and D28.2 gated F2 on SME ratification. D31.5 rules it
# back IN: code-mixed is not a deferred register — it retrieves at 100% for
# every ANSWERABLE question in the set — so a documented refusal that cannot be
# reached in it is an index gap, not a language gap.
#
# WHY A FRAME TABLE AND NOT A TRANSLATION. Nothing here translates. An officer
# typing code-mixed keeps the domain nouns in English — scheme, pension scheme,
# beneficiary, village, cash, purpose — and switches only the INTERROGATIVE
# FRAME around them: "kitne … hain?", "sabse zyada … kis … me?". So the rule
# rewrites the frame and leaves every content word exactly as the workbook wrote
# it. That is deterministic, inspectable in one table, and cannot invent a
# domain term the catalogue does not use.
#
# IT IS APPLIED TO THE SCOPE-FREE FORM, not the raw question, so the place and
# the year are already gone — which is right twice over: the officer states them
# in their own sentence, and "a given Block" has no code-mixed rendering worth
# indexing.
#
# SAFETY. A paraphrase can only ever RAISE the entry it belongs to (an entry
# scores as the MAX over its vectors), so the worst case is an entry that
# out-ranks another — which is the intended effect where it happens and is
# measured for all 30 entries by `refusal_recall.py`, not asserted by argument.

# (pattern, replacement) applied in order to a scope-free question. Every
# replacement keeps the ORIGINAL noun phrase, captured as \1.
_HINGLISH_FRAMES: tuple[tuple["re.Pattern[str]", str], ...] = (
    # "How many beneficiaries received benefits for each stated purpose?"
    (re.compile(r"^How many (.+?) received benefits(.*?)\??$", re.I),
     r"Kitne \1 ko benefit mila\2?"),
    # "How many beneficiaries are recorded under each pension scheme?"
    (re.compile(r"^How many (.+?) (?:are|is) recorded(.*?)\??$", re.I),
     r"Kitne \1 darj hain\2?"),
    # "How many beneficiary records are missing the village field?"
    (re.compile(r"^How many (.+?)\??$", re.I),
     r"Kitne \1 hain?"),
    # "What is the total cash benefit quantum distributed under a given Scheme?"
    (re.compile(r"^What is the total (.+?)\??$", re.I),
     r"Total \1 kitna hai?"),
    # "What is the village-wise count of beneficiaries under a given Scheme?"
    (re.compile(r"^What is the (.+?)\??$", re.I),
     r"\1 kya hai?"),
    # "Which scheme has the most recorded beneficiaries?"
    (re.compile(r"^Which (.+?) has the most (.+?)\??$", re.I),
     r"Sabse zyada \2 kis \1 me hain?"),
    # "Which GPs have no recorded beneficiaries under a given Scheme?"
    (re.compile(r"^Which (.+?) have no (.+?)\??$", re.I),
     r"Kin \1 me koi \2 nahi hai?"),
    (re.compile(r"^Which (.+?)\??$", re.I),
     r"Kaunse \1?"),
    # "List the beneficiaries of a given Scheme with village and benefit details."
    (re.compile(r"^List the (.+?)\.?\??$", re.I),
     r"\1 ki list dikhao"),
    # "Compare beneficiary counts under a given Scheme between …"
    (re.compile(r"^Compare (.+?)\.?\??$", re.I),
     r"\1 compare karo"),
    # "How has the beneficiary count under a given Scheme changed?"
    (re.compile(r"^How has (.+?) changed(.*?)\.?\??$", re.I),
     r"\1 me kya change aaya\2?"),
)

# Residual English function words the frames do not consume. Rewritten AFTER a
# frame matches, never on their own — a sentence that matched no frame is left
# alone rather than half-converted.
_HINGLISH_CONNECTIVES: tuple[tuple["re.Pattern[str]", str], ...] = (
    (re.compile(r"\bunder (?:a given|each) ", re.I), "har "),
    (re.compile(r"\bunder ", re.I), "ke under "),
    (re.compile(r"\bacross GPs\b", re.I), "GPs me"),
    (re.compile(r"\bfor each\b", re.I), "har"),
    (re.compile(r"\bof a given ", re.I), "ke "),
    (re.compile(r"\bwith ", re.I), "ke saath "),
    (re.compile(r"\bbetween ", re.I), "ke beech "),
    (re.compile(r"\ba given ", re.I), ""),
    (re.compile(r"\s{2,}"), " "),
)


def code_mixed_question(question: str) -> str | None:
    """A Hinglish rendering of one refusal question, or None.

    None when no frame matches — a sentence this table does not recognise gets
    no code-mixed line at all, rather than a half-converted one that would
    embed as neither register.
    """
    text = (question or "").strip()
    if not text:
        return None
    for pattern, replacement in _HINGLISH_FRAMES:
        rewritten, count = pattern.subn(replacement, text)
        if not count:
            continue
        for connective, into in _HINGLISH_CONNECTIVES:
            rewritten = connective.sub(into, rewritten)
        rewritten = re.sub(r"\s+([?.])", r"\1", rewritten).strip()
        rewritten = re.sub(r"\s{2,}", " ", rewritten)
        if not rewritten or rewritten.lower() == text.lower():
            return None
        return rewritten[0].upper() + rewritten[1:]
    return None


def build_paraphrases(row: dict, abstract: str, slots: list[dict]) -> list[str]:
    """Extra retrieval surface for one template, deduplicated.

    Three sources, in descending order of how much they help:
      * the Example Question — real values, the shape a user actually types;
      * the Original Question — the officer's own abstract wording, which is
        often phrased quite differently from the parameterised one;
      * one scope line per geography tier this template can filter on (D2).
    """
    out: list[str] = []
    seen = {to_prose(abstract).lower()}

    def add(candidate: str) -> None:
        candidate = re.sub(r"\s+", " ", candidate or "").strip()
        if not candidate or candidate.lower() in seen:
            return
        seen.add(candidate.lower())
        out.append(candidate)

    add(text(row["Example Question"]))
    add(to_prose(text(row["Original Question"])))
    add(to_prose(text(row["Parameterized Question"])))

    optional_geo = [s["name"] for s in slots
                    if s.get("optional") and s["name"] in SCOPE_NOUN]
    if optional_geo:
        source = text(row["Parameterized Question"]) or text(row["Original Question"])
        has_geo_run = bool(_GEO_RUN.search(source))
        for slot in optional_geo:
            if has_geo_run:
                # The question names its own scope — swap that run for this tier
                # so the variant reads as a real question rather than a list of
                # every tier at once.
                add(to_prose(_GEO_RUN.sub(SCOPE_NOUN[slot], source)))
            else:
                # A naturally state-wide question ("how many SWM activities were
                # planned in 2024-25?"), which D2 says the same template must
                # also answer for one district / block / GP.
                add(f"{to_prose(source).rstrip('?. ')}, {SCOPE_SUFFIX[slot]}?")
    return out


# ── Emitting Python ───────────────────────────────────────────────────────────

def py(value, indent: int = 0) -> str:
    """repr() with the quoting style the rest of the catalogue uses."""
    pad = " " * indent
    if isinstance(value, str):
        if "\n" in value:
            # SQL is emitted as a readable triple-quoted block rather than a
            # repr with \n escapes, so a reviewer can read it as SQL. That is
            # only safe while no statement contains a backslash or a triple
            # quote; assert rather than escape, so a future workbook that does
            # stops the build instead of emitting something subtly wrong.
            if "\\" in value or '"""' in value:
                raise ValueError("SQL contains a backslash or triple quote")
            return f'"""\n{value.strip()}\n"""'
        return repr(value)
    if isinstance(value, bool) or value is None or isinstance(value, (int, float)):
        return repr(value)
    if isinstance(value, dict):
        inner = ", ".join(f"{k!r}: {py(v)}" for k, v in value.items())
        return "{" + inner + "}"
    if isinstance(value, list):
        if not value:
            return "[]"
        lines = ",\n".join(f"{pad}    {py(v, indent + 4)}" for v in value)
        return "[\n" + lines + f",\n{pad}]"
    raise TypeError(type(value))


def emit_entry(qid: str, entry: dict) -> str:
    order = ["abstract_question", "date_filter", "date_kind", "sql_template",
             "param_slots", "grouped_geo", "result_ttl_seconds", "caveat",
             "bracket", "module", "submodule", "question_type", "answerable",
             "paraphrases", "notes"]
    lines = [f"    {qid!r}: {{"]
    for key in order:
        if key not in entry:
            continue
        value = entry[key]
        if key == "sql_template":
            lines.append(f'        "sql_template": {py(value)},')
        elif key == "param_slots":
            lines.append('        "param_slots": [')
            for slot in value:
                lines.append(f"            {py(slot)},")
            lines.append("        ],")
        elif key == "paraphrases":
            lines.append('        "paraphrases": [')
            for text_ in value:
                lines.append(f"            {text_!r},")
            lines.append("        ],")
        else:
            lines.append(f'        "{key}": {py(value)},')
    lines.append("    },")
    return "\n".join(lines)


GENERATED_BANNER = '''# ── GENERATED FILE — do not edit by hand ─────────────────────────────────────
# Built from AI_Chatbot_Questions.xlsx by tools/build_catalog.py.
# To change a question, a caveat or a SQL string, change the WORKBOOK and
# regenerate; `python tools/build_catalog.py --check` fails if this file and the
# workbook have drifted apart.
'''

TEMPLATE_HEADER = '''"""
Odisha Panchayati Raj & Drinking Water — the query catalogue.

346 templates: every question on the workbook's Questions sheet that the
database can answer (95 "Yes", 251 "Partial"). The 17 "No" rows and the 13
dropped beneficiary questions are NOT here — they live in
`unanswerable_catalog.py`, so that a question the database genuinely cannot
answer is retrieved and refused with its reason instead of missing.

ONE TEMPLATE PER QUESTION, SLOTS OPTIONAL (decision D2)
    The filter idiom is `($p IS NULL OR col = $p)`: an absent optional slot
    binds SQL NULL, which reads as "do not filter on this". So ONE entry answers
    a question state-wide, for a district, for a block or for a single GP,
    instead of the AP catalogue's -S/-D/-M sibling variants. Absent optional
    slots must never stall on a clarification — "how many activities are
    planned?" is a complete question state-wide.

    Because the scope is no longer visible in the id, it has to be visible to
    RETRIEVAL: each geography-optional entry carries scope-phrased paraphrases
    ("district-wise…", "block-wise…", "GP-wise…") that all resolve to the one
    entry. The vector retriever keeps the MAX score over a template's vectors and
    counts distinct query_ids toward k, so these can never crowd the candidate
    list.

NAMED PARAMETERS (decision D1)
    Every statement uses `$name` placeholders and binds a DICT, one entry per
    slot name however often the name occurs — which the optional-filter idiom
    makes the normal case, since each parameter appears twice. `param_style` is
    sniffed from the SQL by query_router/sql_params.py; no entry sets it.

GEOGRAPHY BINDS A CODE, NOT A NAME (decisions D4/D10)
    Every `$gp_name` slot is `{"bind": "code"}` and binds the resolved
    `gp_lgd_code`. The SQL predicate was rewritten to match: `v.gp_name =
    $gp_name` became `v.gp_lgd_code = $gp_name`. THE SLOT KEEPS ITS NAME AND ITS
    VALUE IS NOW A CODE — the name is the workbook's, the value is the
    validator's. Statewide there are ~6,800 GPs and names repeat freely, so a
    name predicate would silently merge every namesake into one answer.

    `v_asset` carries geography but not `gp_lgd_code`, so its 13 GP predicates
    resolve through the parent activity
    (`activity_code IN (SELECT … FROM v_activity WHERE gp_lgd_code = $gp_name)`),
    which is exactly equivalent. Blocks and districts still bind their validated
    NAME, because no view exposes their codes — see WP3_REPORT for the
    view-change request that would finish the job.

CAVEATS ARE FIRST-CLASS (decision D3)
    296 of these 346 entries carry a `caveat`, verbatim from the workbook's
    Answerability Note. 251 questions are only PARTIALLY answerable — a proxy
    column, a coverage gap, a denominator that is the 20 loaded GPs rather than
    the roster — and a Partial answer served without its caveat is the
    confidently-wrong failure mode this whole layer exists to prevent. The
    caveat reaches the user as `QueryResponse.caveat` AND appended verbatim to
    the rendered answer; it is never passed through an LLM prompt, where it
    could be paraphrased away.

DATES ARE ORDINARY SLOTS (decision D9)
    `date_filter` is None on every entry and `date_kind` is never set. The
    fiscal year is a normal `$date_range` slot binding the full `'2024-2025'`
    string; `date_phrase.py` maps "FY 24-25" / "last year" onto it. The engine's
    date-injection machinery stays dormant for PR&DW — its `year` kind compares
    integers and would raise a binder error on this column.

THE VIEWS
    Every statement reads the `v_*` analytical views, which are created at
    adapter startup from `sql/create_views.sql` into the writable in-memory
    catalog (`db_factory._seed_views`). Seventeen queries also touch a base
    table — `gram_panchayat` mostly, for the LEFT-JOIN-from-the-roster shape
    that keeps zero-activity GPs in the answer, which is the finding a review
    meeting wants.

ENTRY KEYS
    abstract_question   the parameterised question, placeholders renamed to slot
                        names so `.format()` works in suggestions.
    sql_template        the workbook's SQL, verbatim but for the geography
                        rewrite above.
    param_slots         [{name, entity_type, optional?, bind?}] in the
                        workbook's bind order. `entity_type` matches
                        entity_validator.PARAM_ENTITY_TYPES (a test asserts it).
    caveat              the Answerability Note, when there is one.
    bracket / module / submodule / question_type / answerable
                        the workbook's own classification. Bracket+Module+
                        Submodule is the family structure rerank_context.py
                        groups by.
    paraphrases         extra retrieval surface; see D2 above.

VALIDATION
    tests/test_catalog_execution.py executes all 346 against the sample database
    with the workbook's own sample parameters and compares row counts against
    the Test Report sheet. SQL is deterministic, so any mismatch is a real
    defect rather than replay noise.
"""
'''

UNANSWERABLE_HEADER = '''"""
The questions this database CANNOT answer, and why.

Thirty of them: the 17 rows the workbook marks "Answerable from DB = No", and
the 13 beneficiary questions on its Dropped sheet.

WHY THEY ARE IN THE CATALOGUE AT ALL
    Officers will ask these. Beneficiary questions especially — "how many
    beneficiaries got a pension in this GP" is an obvious thing to want, and the
    only beneficiary table in the database is empty. Left out of the catalogue,
    such a question retrieves nothing, scores below the no-match threshold and
    gets the generic "I'm not sure I can answer that specific question yet"
    fallback, which is indistinguishable from the bot merely failing. The
    officer is left unsure whether to rephrase, and a bot that looks unreliable
    on an answerable-sounding question is worse than one that says plainly what
    it does not hold.

    So these are RETRIEVABLE but NOT EXECUTABLE. They carry no SQL and no slots.
    `router` serves a matched id from here as an honest refusal built from the
    workbook's own reason — "the database cannot answer this because …" — and
    offers the nearest answerable questions where the note names one.

    They are a SEPARATE dict from TEMPLATE_CATALOG deliberately. Everything that
    iterates TEMPLATE_CATALOG assumes an entry has SQL and slots — the binder,
    the execution gate, `validate_catalog`, `_accepted_filters`. Merging the two
    would mean teaching every one of those about a third kind of entry; keeping
    them apart means the retriever indexes both and only the router has to know.

KEYS
    question    the workbook's Original Question, placeholders written out.
    reason      the Answerability Note / drop reason, VERBATIM. This is what the
                user is told; it is the whole value of the entry.
    alternative a query_id that answers the nearest answerable question, where
                the workbook's note names one. Offered as a chip, never
                substituted silently.
"""
'''

TEMPLATE_FOOTER = '''

ALL_TEMPLATES: dict[str, dict] = TEMPLATE_CATALOG


def bind(template_id: str, values: dict):
    """Return (sql, params) for a template and a {slot_name: value} dict.

    Every PR&DW template is NAMED, so this returns a DICT with one entry per
    slot name however often that name occurs in the statement — which is the
    normal case here, since `($p IS NULL OR col = $p)` writes each parameter
    twice. The positional branch is kept because `sql_params.param_style` still
    detects positional SQL and the engine must not silently mis-bind it.

    Slots marked {"optional": True} may be absent from `values`; they bind None
    (SQL NULL). Missing REQUIRED slots still raise.
    """
    from .sql_params import NAMED, param_style

    t = ALL_TEMPLATES[template_id]
    slots = t['param_slots']
    optional = {s['name'] for s in slots if s.get('optional')}
    missing = {s['name'] for s in slots} - set(values) - optional
    if missing:
        raise KeyError(f'{template_id} missing slot values: {sorted(missing)}')

    if param_style(t) == NAMED:
        return t['sql_template'], {s['name']: values.get(s['name']) for s in slots}

    ordered = sorted(slots, key=lambda s: s['position'])
    return t['sql_template'], [values.get(s['name']) for s in ordered]


def required_entities(template_id: str) -> list[str]:
    """Distinct entity types the extractor must resolve for this template."""
    return sorted({s['entity_type'] for s in ALL_TEMPLATES[template_id]['param_slots']})


def retrieval_corpus():
    """(text, template_id) pairs for embedding — abstract question plus paraphrases."""
    pairs = []
    for tid, t in ALL_TEMPLATES.items():
        pairs.append((t['abstract_question'], tid))
        for p in t.get('paraphrases', []):
            pairs.append((p, tid))
    return pairs


def family_of(template_id: str) -> tuple[str, str, str]:
    """(bracket, module, submodule) — the grouping rerank_context.py families on."""
    t = ALL_TEMPLATES[template_id]
    return (t['bracket'], t['module'], t['submodule'])
'''


# ── Family descriptions for the reranker ──────────────────────────────────────
#
# A FAMILY here is a maximal set of templates with IDENTICAL SQL and identical
# slots. That is the only grouping under which "siblings share one description
# word-for-word" is true rather than merely tidy: those members execute the same
# statement with the same binds, so which one the reranker picks cannot change
# the answer. There are 14 such groups covering 33 template ids — some are the
# workbook's own scope variants collapsed by D2's optional filters (PLN-025 "in
# {GP}" / PLN-027 "in {Block}" / PLN-029 "in {District}" are one query), and a
# few are duplicate questions.
#
# Everything else gets its OWN description, and that is deliberate. The AP lesson
# — "per-variant descriptions caused confusion, not precision" — is about
# PARAMETER variants, where `accepts filters:` already tells the model which
# sibling to take. It does not transfer here: SBM-SWM-002 (community compost
# pits) and SBM-SWM-007 (household compost pits) accept identical filters and
# differ only in a keyword regex frozen inside the SQL. Giving them one
# description would leave the reranker choosing between them blind.
#
# The descriptions are BUILT FROM THE SQL rather than hand-written prose,
# because what the reranker needs is exactly what the SQL says and the question
# text does not: the measure and its accounting basis, the row grain, the status
# filter, and — for the 85 SBM families — which keywords define the concept. A
# hand-written line would be a less accurate way of saying the same thing, and
# would drift from the SQL the first time the workbook is re-ratified.
# `_DISAMBIGUATION` carries the hand-authored half: the near-miss warnings no
# amount of SQL parsing can infer.

MEASURE_GLOSS = {
    "total_expenditure": "actual expenditure on the PLAN basis (activity_expenditure), "
                         "not the cashbook",
    "actual_expenditure": "actual expenditure on the PLAN basis (activity_expenditure), "
                          "not the cashbook",
    "expenditure": "actual expenditure on the PLAN basis (activity_expenditure), "
                   "not the cashbook",
    "planned_cost": "the planned cost the GP entered in the action plan",
    "approved_cost_action_plan": "the action-plan approved cost",
    "admin_approved_cost": "the administratively approved cost",
    "work_proposed_cost": "the cost proposed in the approval order",
    "tec_approval_cost": "the technically approved cost",
    "fund_sanctioned_total": "funds sanctioned under the approval's scheme components",
    "pct_utilised": "utilisation as a percentage of the approved cost",
    "amount": "CASHBOOK voucher amounts (cash basis), which is a different "
              "convention from the plan-basis expenditure most questions use",
}

STATUS_CLAUSES = [
    (r"\bis_completed\s*=\s*1", "counts only WORK COMPLETED"),
    (r"\bis_ongoing\s*=\s*1", "counts only WORK ONGOING"),
    (r"\bis_abandoned\s*=\s*1", "counts only WORK ABANDONED"),
    (r"\bis_under_approval\s*=\s*1", "counts only activities UNDER APPROVAL"),
    (r"\bis_started\s*=\s*0", "counts only activities NOT YET STARTED"),
    (r"\bis_admin_approved\s*=\s*1", "restricted to activities that HAVE an "
                                     "administrative approval (17% of them)"),
    (r"\bis_approved\s*=\s*1", "restricted to plans with an approval date"),
    (r"\bis_costless_activity\s*=\s*1", "restricted to zero-cost activities"),
    (r"\bhas_progress_evidence\s*=\s*1", "restricted to activities with geotagged "
                                         "progress evidence"),
    (r"COALESCE\(v\.total_expenditure,\s*0\)\s*=\s*0",
     "restricted to activities with NO expenditure recorded"),
    (r"tied_untied\s*=\s*'Tied'", "restricted to TIED grant funds"),
    (r"tied_untied\s*=\s*'Untied'", "restricted to UNTIED grant funds"),
]

# Hand-authored near-miss warnings, appended to the family a query_id belongs to.
# These are the distinctions the SQL cannot state about itself.
_DISAMBIGUATION: dict[str, str] = {
    "PLN-001": "UPLOADED a GPDP, which is any plan row — not the approved subset; "
               "the approval question is the neighbouring one.",
    "PLN-002": "APPROVED plans specifically. Because plan_code_status is entirely "
               "NULL, approval is proxied by an approval date, and every loaded "
               "plan has one, so this currently returns the same figure as the "
               "'uploaded' question — say so rather than presenting them as two "
               "different findings.",
    "PLN-005": "The GPs that filed NOTHING — an absence, listed from the roster by "
               "LEFT JOIN, so a GP with no plan still appears. That is the finding "
               "a review meeting wants and the opposite of the counting questions.",
    "EXP-001": "TOTAL SPEND on the plan basis. The cashbook questions read "
               "v_voucher and answer a different accounting convention; never "
               "substitute one for the other.",
    "TRD-010": "A two-GP head-to-head. Both GPs must be named; a question about one "
               "GP alone belongs to an ordinary GP-filtered family.",
    "TRD-011": "A two-BLOCK head-to-head. Both blocks must be named.",
    "TRD-012": "One district read against the STATE benchmark. Every district is "
               "returned so the chosen one can be seen in context — the district "
               "filter deliberately does not narrow the table.",
    "DQY-001": "A DATA-QUALITY check on how much of the data decodes at all, not a "
               "programme measure. Do not answer a coverage or spend question with it.",
}


def _select_columns(sql: str) -> list[str]:
    """The output column names of the outermost SELECT."""
    body = mask_literals(sql)
    # Skip a leading CTE so the reported columns are the ones the user sees.
    tail = body
    if re.match(r"\s*WITH\b", body, re.IGNORECASE):
        depth = 0
        for i, ch in enumerate(body):
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            elif depth == 0 and body[i:i + 7].upper() == "SELECT " and i > 4:
                tail = body[i:]
                break
    m = re.search(r"\bSELECT\b(.*?)\bFROM\b", tail, re.S | re.IGNORECASE)
    if not m:
        return []
    depth, current, parts = 0, "", []
    for ch in m.group(1):
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        if ch == "," and depth == 0:
            parts.append(current)
            current = ""
        else:
            current += ch
    parts.append(current)

    names = []
    for part in parts:
        part = " ".join(part.split())
        if not part:
            continue
        alias = re.search(r"\bAS\s+(\w+)\s*$", part, re.IGNORECASE)
        names.append(alias.group(1) if alias else part.split(".")[-1])
    return names


GEO_COLUMNS = {"gp_name", "block_name", "district_name", "gp_lgd_code",
               "state_name", "activity_code", "plan_code", "fiscal_year", "item"}

# slot name -> how it reads in prose, the inverse of QUESTION_TOKENS.
SLOT_PROSE = {slot: TOKEN_PROSE[token] for token, slot in QUESTION_TOKENS.items()}


def describe_family(entry: dict, member_ids: list[str]) -> str:
    """One line saying what this family's answer actually is."""
    sql = entry["sql_template"]
    masked = mask_literals(sql)
    columns = _select_columns(sql)
    measures = [c for c in columns if c not in GEO_COLUMNS]

    kind = {
        "Count": "COUNTS", "Aggregation": "TOTALS", "Ranking": "RANKS",
        "Listing": "LISTS", "Rate/Percentage": "The PERCENTAGE for",
        "Trend": "The YEAR-BY-YEAR trend of", "Comparison": "COMPARES",
        "Status Lookup": "LOOKS UP", "Lookup": "LOOKS UP",
    }.get(entry["question_type"], "Reports")

    # Placeholders are written OUT rather than deleted: dropping {district_name}
    # from "…in {district_name}/{block_name} have uploaded…" leaves "…in/ have
    # uploaded…", which reads as damage rather than as a description.
    subject = re.sub(
        r"\{(\w+)\}",
        lambda m: SLOT_PROSE.get(m.group(1), m.group(1).replace("_", " ")),
        entry["abstract_question"],
    )
    subject = re.sub(r"\s+", " ", subject).strip(" ?.,")
    parts = [f"{kind}: {subject}."]

    # A pure listing has no measure columns at all (PLN-005 returns three place
    # names); say what it lists rather than saying nothing.
    shown = measures or columns
    if shown:
        parts.append("Returns " + ", ".join(shown[:6]) + ".")

    grain = re.search(r"\bGROUP BY\b([^\n]*)", masked, re.IGNORECASE)
    group_cols = [c for c in columns[:3] if c in GEO_COLUMNS or c in
                  ("theme", "focus_area_name", "status_label", "work_type_label",
                   "asset_category_label", "funding_source", "sanction_authority")]
    if grain and group_cols:
        parts.append("One row per " + " × ".join(group_cols) + ".")
    elif not grain and entry["question_type"] in ("Count", "Aggregation",
                                                  "Rate/Percentage"):
        parts.append("A single summary row.")
    elif not grain:
        parts.append("One row per matching record.")

    order = re.search(r"\bORDER BY\b[^\n]*?\b(ASC|DESC)\b", masked, re.IGNORECASE)
    if "LIMIT $top_n" in masked:
        direction = "lowest first" if (order and order.group(1).upper() == "ASC") \
            else "highest first"
        parts.append(f"A top-N list, {direction}; $top_n = 1 answers "
                     f"'which is the single highest/lowest'.")

    concept = re.search(r"regexp_matches\(\s*v\.search_text\s*,\s*'([^']*)'", sql)
    if concept:
        # The pattern is quoted VERBATIM rather than split into a word list:
        # several of these use alternation inside groups
        # ('(compost).*(community|group|cluster)'), and splitting on the pipe
        # turns a precise rule into a misleading one.
        parts.append(
            f"Identifies the activity by KEYWORD MATCH on activity_name + "
            f"activity_desc against the pattern /{concept.group(1)}/. Nothing in "
            f"the database codes SBM activity types, so this is a text search: "
            f"it both misses differently-worded activities and picks up "
            f"unrelated ones."
        )

    for pattern, clause in STATUS_CLAUSES:
        if re.search(pattern, masked, re.IGNORECASE):
            parts.append(clause.capitalize() + ".")

    # The bootstrap's load-bearing shape: a question answered FROM THE ROSTER, so
    # that a GP with nothing to show still appears. "Which GPs planned nothing
    # under this theme" cannot be answered by any query over v_activity alone —
    # the rows it wants do not exist there — and an activity-side near neighbour
    # would silently answer a different question.
    if re.search(r"\bFROM\s+gram_panchayat\b", masked, re.IGNORECASE):
        if re.search(r"\bNOT\s+EXISTS\b", masked, re.IGNORECASE):
            parts.append(
                "An ABSENCE, read from the GP ROSTER: the Gram Panchayats with no "
                "matching record at all. A query over the activity table alone can "
                "never return these rows, because the rows it would need do not "
                "exist there."
            )
        elif re.search(r"\bLEFT\s+JOIN\b", masked, re.IGNORECASE):
            parts.append(
                "Counted FROM THE GP ROSTER by LEFT JOIN, so panchayats with zero "
                "activity are still in the denominator and still appear — which is "
                "the finding a review meeting is looking for."
            )

    for measure in measures:
        gloss = MEASURE_GLOSS.get(measure)
        if gloss:
            parts.append(f"{measure} is {gloss}.")
            break

    if re.search(r"\bv_voucher\b", masked):
        parts.append("Reads the CASHBOOK (v_voucher) — cash basis, a different "
                     "convention from the plan-basis expenditure questions.")

    geo = [s["name"] for s in entry["param_slots"]
           if s.get("optional") and s["name"] in SCOPE_NOUN]
    if geo:
        tiers = ", ".join(n.replace("_name", "").replace("gp", "GP") for n in geo)
        parts.append(f"Filterable by {tiers}; answers state-wide when no place is "
                     f"named — one entry serves every scope.")

    for qid in member_ids:
        if qid in _DISAMBIGUATION:
            parts.append(_DISAMBIGUATION[qid])
            break

    caveat = (entry.get("caveat") or "").strip()
    if caveat:
        first = re.split(r"(?<=[.;])\s", caveat)[0]
        parts.append("Caveat: " + first.rstrip(".") + ".")

    return " ".join(" ".join(p.split()) for p in parts)


RERANK_HEADER = '''"""
Family descriptions for the re-ranker's "↳" line.

The re-ranker shows each candidate as

    EXP-001: What is the total actual expenditure incurred by {gp_name} in {date_range}?
        ↳ <desc>
        accepts filters: date_range, district_name, block_name, gp_name

and its system prompt tells the model to judge a candidate by the ↳ description
rather than by surface word overlap. This module supplies those descriptions.

WHAT A FAMILY IS HERE
    A maximal set of templates with IDENTICAL SQL and identical slots — the only
    grouping under which the contract "siblings repeat one description
    word-for-word" is true rather than merely tidy, because those members
    execute the same statement and the reranker's choice between them cannot
    change the answer. 14 such groups cover 33 of the 346 ids; the workbook's
    own scope variants are most of them (PLN-025 "in {GP}", PLN-027 "in
    {Block}", PLN-029 "in {District}" are one query once D2 makes geography
    optional). The other 313 templates are each their own family.

    That is a deliberate departure from the AP catalogue, where descriptions
    covered large families. The AP lesson — per-variant descriptions caused
    confusion, not precision — is about PARAMETER variants, which the
    "accepts filters:" line already separates. It does not transfer to this
    catalogue: SBM-SWM-002 (community compost pits) and SBM-SWM-007 (household
    compost pits) accept identical filters and differ only in a keyword regex
    frozen inside the SQL, so one shared description would leave the model
    choosing between them blind.

WHAT A DESCRIPTION SAYS, and why it is generated
    Descriptions are BUILT FROM THE SQL by tools/build_catalog.py, because what
    the reranker is missing is exactly what the SQL knows and the question text
    does not:

      1. the measure AND its accounting basis — the data dictionary's central
         trap is that this database holds two expenditure conventions, plan
         basis (activity_expenditure) and cash basis (the voucher cashbook), and
         an answer that does not say which is a wrong answer;
      2. the row grain — one row per GP, per theme, or a single total;
      3. the status filter — "only WORK COMPLETED", "only activities with an
         administrative approval", "only those with no expenditure recorded";
      4. for the 85 SBM families, WHICH KEYWORDS define the concept, since
         nothing in the database codes SBM activity types and every one of those
         questions is a text search;
      5. the scope behaviour — one entry answers state-wide or narrowed.

    Hand-written prose would be a less accurate way of saying the same things
    and would drift from the SQL at the first re-ratification. The hand-authored
    half is `_DISAMBIGUATION` in the builder: the near-miss warnings no amount of
    SQL parsing can infer ("uploaded a GPDP is any plan row, not the approved
    subset").

CONTRACT (mirrors _RERANK_SYS in reranker.py — do not break it)
    - ONE line, no newlines: the candidate listing is line-oriented.
    - Every template has a non-empty description, and no template appears in two
      families. tests/test_rerank_context.py enforces both.
"""
'''


def build_rerank_context(sheets: dict, templates_src: str) -> tuple[str, int]:
    """Emitted after the template catalogue, from the catalogue itself."""
    # The generated module has no top-level imports — `bind()` imports
    # sql_params inside its body — so it executes standalone, and reading the
    # catalogue back out of it is the surest way for the descriptions to be
    # about the entries that were actually written.
    namespace: dict = {}
    exec(compile(templates_src, "<template_catalog>", "exec"), namespace)
    catalog = namespace["TEMPLATE_CATALOG"]

    families: dict[tuple, list[str]] = defaultdict(list)
    for qid, entry in catalog.items():
        key = (" ".join(entry["sql_template"].split()),
               tuple((s["name"], s.get("optional", False))
                     for s in entry["param_slots"]))
        families[key].append(qid)

    blocks = []
    for members in sorted(families.values(), key=lambda m: sorted(m)[0]):
        members = sorted(members)
        entry = catalog[members[0]]
        name = _family_name(members, entry)
        desc = describe_family(entry, members)
        assert "\n" not in desc
        blocks.append(
            f"    {name!r}: {{\n"
            f"        \"desc\": {desc!r},\n"
            f"        \"members\": {members!r},\n"
            f"    }},"
        )

    body = (
        RERANK_HEADER
        + GENERATED_BANNER
        + "\n\nFAMILY_DESCRIPTIONS: dict[str, dict] = {\n\n"
        + "\n\n".join(blocks)
        + "\n}\n\n"
        + '''
# query_id -> family description, expanded from the members lists above.
DESC_BY_QID: dict[str, str] = {
    qid: family["desc"]
    for family in FAMILY_DESCRIPTIONS.values()
    for qid in family["members"]
}
'''
    )
    return body, len(families)


def _family_name(members: list[str], entry: dict) -> str:
    """A readable key: the lead id plus a slug of the submodule."""
    slug = re.sub(r"[^a-z0-9]+", "_", entry["submodule"].lower()).strip("_")
    return f"{members[0].lower().replace('-', '_')}__{slug}"


# ── Build ─────────────────────────────────────────────────────────────────────

def build_templates(sheets: dict) -> tuple[str, list[str], dict]:
    rows = [r for r in sheets["Questions"]
            if text(r["Answerable from DB"]) in ("Yes", "Partial")]
    registry_optional = {
        text(r["Bind name"]).lstrip("$"): text(r["NULL skips filter"]) == "Yes"
        for r in sheets["Parameter Registry"]
    }

    findings: list[str] = []
    entries: list[str] = []
    stats = Counter()
    disagreements: list[str] = []
    asset_rewrites: list[str] = []

    for row in rows:
        qid = text(row["Question ID"])
        sql = text(row["Parameterized SQL"])
        if not sql:
            raise ValueError(f"{qid}: answerable but has no SQL")

        sql, notes = rewrite_geography(qid, sql)
        if any("v_asset" in n for n in notes):
            asset_rewrites.append(qid)
        stats["rewritten"] += 1 if notes or "gp_lgd_code = $gp_name" in sql else 0

        slots = build_slots(row, sql)
        slot_names = {s["name"] for s in slots}
        abstract = to_abstract(text(row["Parameterized Question"]), slot_names, qid)

        for slot in slots:
            name = slot["name"]
            if name in DEFAULTED_SLOTS:
                # D18.P1 overrides both sources for these; counting them as
                # disagreements would bury the 12 real ones under 91 noisy rows.
                stats["defaulted"] += 1
                continue
            declared = registry_optional.get(name)
            if declared is not None and declared != bool(slot.get("optional")):
                disagreements.append(
                    f"{qid} ${name}: SQL guard says "
                    f"{'optional' if slot.get('optional') else 'required'}, "
                    f"Parameter Registry says "
                    f"{'optional' if declared else 'required'}"
                )

        entry = {
            "abstract_question": abstract,
            "date_filter": None,        # decision D9 — never a date_filter
            "date_kind": None,
            "sql_template": sql,
            "param_slots": slots,
            "result_ttl_seconds": 600,
            "bracket": text(row["Bracket"]),
            "module": text(row["Module"]),
            "submodule": text(row["Submodule"]),
            "question_type": text(row["Question Type"]),
            "answerable": text(row["Answerable from DB"]),
            "paraphrases": build_paraphrases(row, abstract, slots),
        }
        grouped = grouped_geo_slots(sql, slots)
        if grouped:
            entry["grouped_geo"] = grouped
            stats["grouped_geo"] += 1
        note = text(row["Answerability Note"])
        if note:
            entry["caveat"] = note
            stats["caveated"] += 1
        stats["templates"] += 1
        entries.append(emit_entry(qid, entry))

    body = (
        TEMPLATE_HEADER
        + GENERATED_BANNER
        + "\n\nTEMPLATE_CATALOG: dict[str, dict] = {\n\n"
        + "\n\n".join(entries)
        + "\n}\n"
        + TEMPLATE_FOOTER
    )

    findings.append(f"templates: {stats['templates']}, caveated: {stats['caveated']}")
    findings.append(
        f"v_asset GP predicates resolved through v_activity ({len(asset_rewrites)}): "
        + ", ".join(asset_rewrites)
    )
    findings.append(
        f"optional-slot disagreements with the Parameter Registry "
        f"({len(disagreements)}) — the SQL guard wins:\n    "
        + "\n    ".join(disagreements)
    )
    findings.append(
        f"optional-with-default slots (D18.P1, {stats['defaulted']} occurrences): "
        + ", ".join(f"${n} = {v!r}" for n, v in sorted(DEFAULTED_SLOTS.items()))
    )
    findings.append(
        f"templates reporting one row per geography (grouped_geo): "
        f"{stats['grouped_geo']}"
    )
    return body, findings, stats


def _unanswerable_paraphrases(
    question: str, extra: list[str], *, code_mixed: bool = False
) -> list[str]:
    """Index surface for one unanswerable entry, deduplicated, order stable.

    The workbook's own alternative wordings first where it has any, then the
    scope-free line (see `scope_free_question` for why it exists and what it
    measured), then the code-mixed line where it is asked for.

    `code_mixed` IS GATED TO THE DROPPED SHEET, which is the scope D31.5 rules
    on: the beneficiary refusals are the entries whose gold questions are typed
    in that register, and BEN-003 is the one measured stuck outside the window
    because of it. The frame table itself is general and would fire on most of
    the 17 "No" rows — on several of them producing a line that differs from the
    English by one word, which is index weight for no retrieval gain. Widening
    the gate is a one-line change and should be made the way this one was: on a
    rank measurement from `refusal_recall.py`, not on the argument that more
    surface cannot hurt.
    """
    out: list[str] = []
    # The code-mixed line is built from the SCOPE-FREE form, so the place and
    # the year are already gone before the frame is switched (D31.5).
    scope_free = scope_free_question(question)
    candidates = [*extra, scope_free]
    if code_mixed:
        candidates.append(code_mixed_question(scope_free or question))
    for candidate in candidates:
        if not candidate:
            continue
        if candidate.strip().lower() == (question or "").strip().lower():
            continue
        if any(candidate.strip().lower() == seen.strip().lower() for seen in out):
            continue
        out.append(candidate.strip())
    return out


def build_unanswerable(sheets: dict) -> tuple[str, int]:
    """The 17 'No' rows plus the 13 Dropped beneficiary questions."""
    entries: list[str] = []
    count = 0

    # A workbook note that names its own answerable alternative, e.g. PLN-022's
    # "PLN-010 with a user-supplied $deadline is the closest answerable form."
    ref = re.compile(r"\b([A-Z]{3}(?:-[A-Z]{2,3})?-\d{3})\b")

    for row in sheets["Questions"]:
        if text(row["Answerable from DB"]) != "No":
            continue
        qid = text(row["Question ID"])
        reason = text(row["Answerability Note"])
        alternatives = [m for m in ref.findall(reason) if m != qid]
        entry = {
            "question": to_prose(text(row["Original Question"])),
            "reason": reason,
            "bracket": text(row["Bracket"]),
            "module": text(row["Module"]),
            "submodule": text(row["Submodule"]),
            "source": "No",
        }
        if alternatives:
            entry["alternative"] = alternatives[0]
        entry["paraphrases"] = _unanswerable_paraphrases(
            entry["question"],
            [to_prose(text(row["Parameterized Question"])),
             text(row["Example Question"])])
        entries.append(emit_unanswerable(qid, entry))
        count += 1

    for row in sheets["Dropped"]:
        qid = text(row["Question ID"])
        entry = {
            "question": to_prose(text(row["Original Question"])),
            "reason": text(row["Why it was dropped"]),
            "bracket": text(row["Bracket"]),
            "module": text(row["Module"]),
            "submodule": text(row["Module"]),
            "source": "Dropped",
            # The Dropped sheet has five columns and none of them is a
            # Parameterized or Example question, so the scope-free line is the
            # only extra surface available — and it is the one that mattered.
            "paraphrases": _unanswerable_paraphrases(
                to_prose(text(row["Original Question"])), [], code_mixed=True),
        }
        entries.append(emit_unanswerable(qid, entry))
        count += 1

    body = (
        UNANSWERABLE_HEADER
        + GENERATED_BANNER
        + "\n\nUNANSWERABLE_CATALOG: dict[str, dict] = {\n\n"
        + "\n\n".join(entries)
        + "\n}\n"
        + '''

def retrieval_corpus():
    """(text, query_id) pairs for embedding, same shape as the template catalogue."""
    pairs = []
    for qid, entry in UNANSWERABLE_CATALOG.items():
        pairs.append((entry["question"], qid))
        for p in entry.get("paraphrases", []):
            pairs.append((p, qid))
    return pairs


def refusal_for(query_id: str) -> str | None:
    """The sentence an officer is shown, built from the workbook's own reason.

    Verbatim, for the same reason a caveat is verbatim: the reason IS the
    answer to "why not", and a paraphrase of it is a worse answer.
    """
    entry = UNANSWERABLE_CATALOG.get(query_id)
    if entry is None:
        return None
    return (
        f"I can't answer that from this database. {entry['reason']}"
    )
'''
    )
    return body, count


def emit_unanswerable(qid: str, entry: dict) -> str:
    lines = [f"    {qid!r}: {{"]
    for key in ("question", "reason", "bracket", "module", "submodule",
                "source", "alternative"):
        if key in entry:
            lines.append(f'        "{key}": {entry[key]!r},')
    lines.append('        "paraphrases": [')
    for p in entry["paraphrases"]:
        lines.append(f"            {p!r},")
    lines.append("        ],")
    lines.append("    },")
    return "\n".join(lines)


def build_oracle(sheets: dict) -> tuple[str, int]:
    """The Test Report sheet as JSON: the row counts the execution gate checks."""
    oracle = {}
    for row in sheets["Test Report"]:
        qid = text(row["Question ID"])
        if text(row["Answerable"]) == "No":
            continue
        params = text(row["Parameters used"])
        oracle[qid] = {
            "answerable": text(row["Answerable"]),
            "status": text(row["Test Status"]),
            "rows": int(row["Rows"]),
            "params": json.loads(params) if params else {},
        }
    return json.dumps(oracle, indent=1, sort_keys=True) + "\n", len(oracle)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    ap.add_argument("--check", action="store_true",
                    help="report drift and write nothing")
    args = ap.parse_args()

    if not args.workbook.exists():
        print(f"workbook not found: {args.workbook}", file=sys.stderr)
        return 2
    lock = args.workbook.parent / ("~$" + args.workbook.name)
    if lock.exists():
        print(f"{args.workbook.name} is open in Excel ({lock.name} present). "
              "Close it so the saved file is current.", file=sys.stderr)
        return 2

    sheets = read_workbook(args.workbook)
    template_src, findings, _ = build_templates(sheets)
    unanswerable_src, n_unanswerable = build_unanswerable(sheets)
    oracle_src, n_oracle = build_oracle(sheets)
    rerank_src, n_families = build_rerank_context(sheets, template_src)

    outputs = [
        (TEMPLATE_OUT, template_src),
        (UNANSWERABLE_OUT, unanswerable_src),
        (RERANK_OUT, rerank_src),
        (ORACLE_OUT, oracle_src),
    ]

    drifted = []
    for path, source in outputs:
        current = path.read_text(encoding="utf-8") if path.exists() else None
        if current != source:
            drifted.append(path)

    for line in findings:
        print(line)
    print(f"unanswerable: {n_unanswerable}   oracle rows: {n_oracle}")

    if args.check:
        if drifted:
            print("\nDRIFT — these do not match the workbook:", file=sys.stderr)
            for path in drifted:
                print(f"  {path.relative_to(BACKEND)}", file=sys.stderr)
            return 1
        print("\nin step with the workbook")
        return 0

    for path, source in outputs:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(source, encoding="utf-8")
        print(f"wrote {path.relative_to(BACKEND)}  ({len(source):,} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
