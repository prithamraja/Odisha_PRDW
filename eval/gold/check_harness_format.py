# -*- coding: utf-8 -*-
"""WP-4a gate evidence: load the gold set with the HARNESSES' OWN parsers.

Run it from the repo:  python eval/gold/check_harness_format.py
Requires the built artefacts, so run the builder first:
    python eval/gold/build_eval_questions.py

Import-level only — no eval run, no LLM calls, no server, no database. Nothing
in the repo is modified.

Three checks:
  1. run_full_eval.py's spec reader — the exact json.loads + .get() field
     accesses it performs per question, including the "session": "prev" path.
  2. grade_full_eval.py's grade() — the real grading function, imported from
     Chatbot/, run over synthetic responses for every gold record.
  3. recall_eval.py's load_gold() — the real CSV reader, pointed at the built
     test_questions_query_mapping.csv.

Checks 2 and 3 import the Chatbot/query_router catalogues. A failure there is
reported as SOFT rather than fatal, so a catalogue mid-rewrite does not read as
a broken gold set. Route- and slot-level cross-checking against the catalogue
lives in build_eval_questions.py --check; run both.
"""
import csv
import io
import json
import os
import shutil
import sys
import tempfile
from pathlib import Path
from uuid import uuid4

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

GOLD = Path(__file__).resolve().parent      # eval/gold/
# Repo root is two levels up. PRDW_REPO overrides it, which is how this ran
# during WP-4a: the gold tree was staged outside the repo (WP-3 owned the
# working tree) while Chatbot/ and the workbook were read from the repo itself.
REPO = Path(os.environ.get("PRDW_REPO") or GOLD.parent.parent)

# THE BACKEND DIRECTORY, RESOLVED RATHER THAN NAMED (WP-5). It was `Chatbot/`
# through WP-4c and is `Ask/` since the tree was reorganised on 2026-08-18 — a
# rename that silently broke both gold-set checks, because a hard-coded name
# does not fail loudly, it just stops finding the catalogue and reports the
# cross-check as SKIPPED. Both spellings are tried, newest first, and an
# unresolvable one is named in the error rather than guessed at.
_BACKEND_DIR_NAMES = ("Ask", "Chatbot")


def _backend_dir(repo):
    for name in _BACKEND_DIR_NAMES:
        candidate = repo / name
        if (candidate / "query_router").is_dir():
            return candidate
    # Nothing found: return the first spelling so the error a caller raises
    # names a real path rather than None.
    return repo / _BACKEND_DIR_NAMES[0]


CHATBOT = _backend_dir(REPO)

def _find(name: str, installed: Path) -> Path:
    """Prefer the installed copy (what the harnesses actually read); fall back to
    the one the builder writes beside the gold files."""
    for cand in (installed, GOLD / name):
        if cand.exists():
            return cand
    sys.exit(f"{name} not found — run build_eval_questions.py first "
             f"(looked in {installed.parent} and {GOLD})")


FULL = _find("eval_questions_full.json", CHATBOT / "eval_questions_full.json")
CSV_PATH = _find("test_questions_query_mapping.csv",
                 REPO / "test_questions_query_mapping.csv")

ok = True
soft = []


def head(t):
    print(f"\n{'='*72}\n{t}\n{'='*72}")


# ── 1. run_full_eval.py's spec reader, replayed verbatim ─────────────────────
head("1. run_full_eval.py spec reader (verbatim field accesses)")

QUESTIONS = json.loads(FULL.read_text(encoding="utf-8"))
print(f"json.loads -> {len(QUESTIONS)} specs")

prev_session = None
n_prev_sessions = 0
seen_n = set()
for i, spec in enumerate(QUESTIONS, 1):
    # exactly what run_full_eval.run() does per spec
    if spec.get("session") == "prev" and prev_session:
        sid = prev_session
        payload = {"message": spec["q"], "session_id": sid}
        n_prev_sessions += 1
    else:
        sid = str(uuid4())
        payload = {"message": spec["q"], "session_id": sid, "reset_context": True}
    prev_session = sid

    rec = {"n": spec["n"], "q": spec["q"], "gold": spec.get("gold"),
           "acc": spec.get("acc", []), "partial": spec.get("partial", False),
           "excluded": spec.get("excluded", False), "src": spec.get("src")}
    assert isinstance(payload["message"], str) and payload["message"].strip()
    assert rec["n"] not in seen_n, f"duplicate n {rec['n']}"
    seen_n.add(rec["n"])
    # the record is written with json.dumps(..., ensure_ascii=False) — prove the
    # Odia-script rows survive that round trip
    json.loads(json.dumps(rec, ensure_ascii=False, default=str))

print(f"all {len(QUESTIONS)} specs consumed; {n_prev_sessions} resolved a prev session")
print("PASS")

# ── 2. grade_full_eval.grade() — the harness's real grading function ─────────
head("2. grade_full_eval.grade() over synthetic responses")

tmp = Path(tempfile.mkdtemp(prefix="wp4a_gate_"))
try:
    sys.path.insert(0, str(CHATBOT))
    os.environ.setdefault("PRDW_LIVE_EXTRACTION", "0")
    import grade_full_eval  # noqa: E402

    def response_for(spec, behave):
        """Synthesise the /query response shape the gold row says is CORRECT."""
        base = {"n": spec["n"], "q": spec["q"], "gold": spec.get("gold"),
                "acc": spec.get("acc", []), "partial": spec.get("partial", False),
                "excluded": spec.get("excluded", False)}
        if behave in ("answer", "context_preserving_reroute"):
            base.update(tier="tier2", query_id=spec["gold"], n_rows=7,
                        clarification=None)
        elif behave == "clarify":
            base.update(tier="clarify", query_id=None, n_rows=None,
                        clarification={"reason": "missing_parameter",
                                       "prompt": "Which year?", "options": []})
        elif behave == "cannot_answer":
            # WP-4 T3: the correct outcome for these 19 is a SERVED REFUSAL —
            # the router retrieved the known-unanswerable entry and returned its
            # query_id with the workbook's own reason. `result` stays None, so
            # n_rows is None; an empty LIST here is the coupling that would flip
            # every one of them to a failure, and grade() now says so out loud
            # instead of mis-bucketing them.
            base.update(tier="fallback", query_id=spec["gold"], n_rows=None,
                        clarification=None)
        else:   # graceful_fallback (out-of-domain) -> gold is "no_match"
            base.update(tier="fallback", query_id=None, n_rows=None,
                        clarification=None)
        return base

    buckets = {}
    for spec in QUESTIONS:
        v = grade_full_eval.grade(response_for(spec, spec.get("expected_behavior")))
        buckets.setdefault(v, []).append(spec)

    for k in sorted(buckets):
        print(f"  {k:<24} {len(buckets[k])}")

    # A gold row whose OWN stated correct behaviour does not grade as a pass is
    # an encoding bug in the gold set, not a router result.
    PASSING = {"hit", "partial", "clarify_gold_offered"}
    bad = [s for v, rows in buckets.items() if v not in PASSING for s in rows]
    if bad:
        # clarify rows have no chips in this synthetic response, so "clarify" is
        # the expected bucket for them — separate that from real problems
        real = [s for s in bad if s.get("expected_behavior") != "clarify"]
        clar = [s for s in bad if s.get("expected_behavior") == "clarify"]
        if clar:
            print(f"\n  note: {len(clar)} ambiguity rows land in 'clarify' because the "
                  f"synthetic response carries no chips;")
            print( "        with the gold id offered as a chip they grade "
                   "'partial' / 'clarify_gold_offered'.")
        if real:
            ok = False
            print(f"\n  FAIL — {len(real)} rows do not grade as a pass under their own "
                  f"stated behaviour:")
            for s in real[:10]:
                print(f"    {s['id']} {s.get('expected_behavior')} {s['q'][:50]}")
    print("PASS" if ok else "FAIL")

    # 2b. chip-carrying variant for the ambiguity rows
    print("\n2b. ambiguity rows re-graded with the gold id offered as a chip")
    import query_router.template_catalog as tc  # noqa: E402
    print(f"    TEMPLATE_CATALOG holds {len(tc.TEMPLATE_CATALOG)} entries. "
          f"grade_full_eval builds Q_TO_ID from")
    print( "    abstract_question text, which a real chip label may not match verbatim,")
    print( "    so inject the workbook's own question text for the gold ids to exercise")
    print( "    chip_ids() -> grade() end to end.")

    import openpyxl  # noqa: E402
    wb = openpyxl.load_workbook(REPO / "AI_Chatbot_Questions.xlsx",
                                read_only=True, data_only=True)
    ws = wb["Questions"]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    qtext = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(hdr, row))
        pq = (rec.get("Parameterized Question") or "").strip()
        if rec.get("Question ID") and pq:
            qtext[rec["Question ID"]] = pq
    for qid, text in qtext.items():
        grade_full_eval.Q_TO_ID.setdefault(text.lower(), qid)

    amb = [s for s in QUESTIONS if s.get("expected_behavior") == "clarify"]
    amb_buckets = {}
    unlabelled = []
    for spec in amb:
        chips = [qtext[q] for q in [spec["gold"], *spec.get("acc", [])] if q in qtext]
        if not chips:
            unlabelled.append(spec)
            continue
        rec = {"n": spec["n"], "q": spec["q"], "gold": spec["gold"],
               "acc": spec.get("acc", []), "partial": spec.get("partial", False),
               "excluded": False, "tier": "clarify", "query_id": None, "n_rows": None,
               "clarification": {"reason": "missing_parameter",
                                 "prompt": "Which year did you mean?",
                                 "options": [{"label": c} for c in chips]}}
        amb_buckets.setdefault(grade_full_eval.grade(rec), []).append(spec)

    for k in sorted(amb_buckets):
        print(f"      {k:<24} {len(amb_buckets[k])}")
    if unlabelled:
        print(f"      (skipped {len(unlabelled)}: gold id has no parameterised text — "
              f"a workbook 'No' row)")
    bad2 = [s for v, rows in amb_buckets.items()
            if v not in {"partial", "clarify_gold_offered"} for s in rows]
    if bad2:
        ok = False
        print(f"      FAIL — {len(bad2)} ambiguity rows did not grade as a pass "
              f"with the gold chip offered")
        for s in bad2[:10]:
            print(f"        {s['id']} {s['q'][:50]}")
    else:
        print("      PASS — every ambiguity row grades as a pass when the clarification "
              "offers its gold chip")

except Exception as exc:  # noqa: BLE001
    soft.append(f"grade_full_eval import/run: {type(exc).__name__}: {exc}")
    print(f"SOFT FAIL (WP-3 owns Chatbot/ right now): {type(exc).__name__}: {exc}")
finally:
    shutil.rmtree(tmp, ignore_errors=True)

# ── 3. recall_eval.load_gold() — the harness's real CSV reader ───────────────
head("3. recall_eval.py load_gold() CSV reader")

# load_gold() is hardwired to HERE.parent/'test_questions_query_mapping.csv';
# replay its body against the built file rather than move repo files around.
rows = []
with open(CSV_PATH, encoding="utf-8-sig") as f:
    for rec in csv.DictReader(f):
        gold = (rec.get("query_code") or "").strip()
        topic = (rec.get("topic") or "").strip()
        if gold and topic:
            rows.append({"topic": topic, "gold": gold,
                         "intent": (rec.get("intent") or "").strip()})

built = sum(1 for _ in CSV_PATH.read_text(encoding="utf-8-sig").splitlines()) - 1
print(f"DictReader -> {len(rows)} scorable rows (file has {built} data lines)")
if len(rows) != built:
    ok = False
    print("FAIL — rows dropped by the reader's own filter")
else:
    print("PASS")

nonascii = [r for r in rows if any(ord(c) > 127 for c in r["topic"])]
print(f"  {len(nonascii)} rows carry non-ASCII text (Odia script) and survived "
      f"the utf-8-sig read")

head("SUMMARY")
print("hard checks:", "PASS" if ok else "FAIL")
if soft:
    print("soft failures (re-run after WP-3 lands):")
    for s in soft:
        print(f"  - {s}")
else:
    print("soft failures: none")
raise SystemExit(0 if ok else 1)
